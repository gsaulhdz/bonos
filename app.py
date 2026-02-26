from flask import Flask, render_template, request, jsonify, send_file
import pg8000.native
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import io
from urllib.parse import urlparse

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_DIR
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

def get_db():
    raw_url = os.environ.get('DATABASE_URL', '')
    if not raw_url:
        raise Exception('DATABASE_URL no esta configurada')
    db_url = raw_url.replace('postgres://', 'postgresql://', 1)
    url = urlparse(db_url)
    conn = pg8000.native.Connection(
        user=url.username,
        password=url.password,
        host=url.hostname,
        port=int(url.port) if url.port else 5432,
        database=url.path.lstrip('/'),
        ssl_context=True
    )
    return conn

def run_query(conn, sql, params=None, fetch=True):
    try:
        if params:
            result = conn.run(sql, *params) if not isinstance(params, dict) else conn.run(sql, **params)
        else:
            result = conn.run(sql)
        if fetch:
            cols = [c['name'] for c in conn.columns]
            return [dict(zip(cols, row)) for row in result]
        return None
    except Exception as e:
        raise e

def init_db():
    conn = get_db()
    conn.run('''CREATE TABLE IF NOT EXISTS trabajadores (
        id SERIAL PRIMARY KEY, nomina TEXT UNIQUE, nombre TEXT,
        sucursal TEXT, puesto TEXT, area TEXT, nombre_suc TEXT)''')
    conn.run('''CREATE TABLE IF NOT EXISTS periodos (
        id SERIAL PRIMARY KEY, nombre TEXT UNIQUE, mes TEXT,
        anio INTEGER, fecha_inicio TEXT, fecha_fin TEXT)''')
    conn.run('''CREATE TABLE IF NOT EXISTS checklists (
        id SERIAL PRIMARY KEY, periodo_id INTEGER, fecha TEXT,
        sucursal TEXT, area TEXT, calificacion REAL, supervisor TEXT)''')
    conn.run('''CREATE TABLE IF NOT EXISTS afectaciones (
        id SERIAL PRIMARY KEY, periodo_id INTEGER, folio TEXT,
        sucursal TEXT, puesto TEXT, nomina TEXT, nombre TEXT,
        incidencia TEXT, fecha TEXT, porcentaje REAL, observacion TEXT)''')
    conn.run('''CREATE TABLE IF NOT EXISTS actas (
        id SERIAL PRIMARY KEY, periodo_id INTEGER, anio INTEGER,
        mes TEXT, almacen TEXT, area TEXT, puesto TEXT, nombre TEXT,
        fecha TEXT, procedimiento TEXT, folio TEXT, observaciones TEXT,
        nomina TEXT, porcentaje_afectacion REAL)''')
    conn.run('''CREATE TABLE IF NOT EXISTS bono_rotulos (
        id SERIAL PRIMARY KEY, periodo_id INTEGER, sucursal TEXT,
        material_pop REAL, limpieza_visual REAL, radio_dpp REAL,
        chequeo REAL, evidencias REAL, total REAL)''')
    conn.close()

def normalizar_sucursal(suc):
    if suc is None: return ''
    return str(suc).strip().lstrip('0').strip()

def detectar_area(puesto):
    puesto = (puesto or '').upper()
    if 'ROTUL' in puesto: return 'ROTULOS'
    elif 'RECIBO' in puesto or 'PROVEEDOR' in puesto: return 'RECIBO'
    else: return 'MESA DE CONTROL'

def calcular_afectacion_acta(observaciones):
    texto = (observaciones or '').upper()
    if 'REFERENCIA' in texto or 'REFERENCIAS' in texto: return -5.0
    for p in ['MAL APLICADO','MALA APLICACIÓN','MALA APLICACION','MOVIMIENTO',
              'FAMILIA DISTINTA','CONVERSION','CONVERSIÓN','NEGATIVO','TIEMPO Y FORMA','EN TIEMPO']:
        if p in texto: return -20.0
    return -5.0

# =================== CARGA DE CATÁLOGO ===================

def cargar_catalogo(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Hoja1"]
    conn = get_db()
    insertados = actualizados = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        nomina = str(int(row[0]))
        nombre = str(row[1]).strip() if row[1] else ''
        puesto = str(row[2]).strip() if row[2] else ''
        num_suc = str(row[4]).strip().lstrip('0').strip() if row[4] else ''
        nombre_suc = str(row[5]).strip() if row[5] else ''
        area = detectar_area(puesto)
        existing = conn.run('SELECT id FROM trabajadores WHERE nomina=:n', n=nomina)
        if existing:
            conn.run('UPDATE trabajadores SET nombre=:nombre, sucursal=:suc, puesto=:puesto, area=:area, nombre_suc=:nsuc WHERE nomina=:nomina',
                     nombre=nombre, suc=num_suc, puesto=puesto, area=area, nsuc=nombre_suc, nomina=nomina)
            actualizados += 1
        else:
            conn.run('INSERT INTO trabajadores (nomina, nombre, sucursal, puesto, area, nombre_suc) VALUES (:nomina,:nombre,:suc,:puesto,:area,:nsuc)',
                     nomina=nomina, nombre=nombre, suc=num_suc, puesto=puesto, area=area, nsuc=nombre_suc)
            insertados += 1
    conn.close()
    return insertados, actualizados

# =================== PROCESAMIENTO ===================

def procesar_checklist(filepath, periodo_id):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    registros = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Fecha': header_found = True; continue
        if not header_found or row[0] is None: continue
        fecha, _, _, _, suc, supervisor, checklist, calificacion, mes, anio = row[:10]
        registros.append({'fecha': str(fecha), 'sucursal': normalizar_sucursal(suc),
                          'area': str(checklist).upper() if checklist else '',
                          'calificacion': float(calificacion) if calificacion else 0,
                          'supervisor': str(supervisor) if supervisor else ''})
    conn = get_db()
    conn.run('DELETE FROM checklists WHERE periodo_id=:pid', pid=int(periodo_id))
    for r in registros:
        conn.run('INSERT INTO checklists (periodo_id, fecha, sucursal, area, calificacion, supervisor) VALUES (:pid,:fecha,:suc,:area,:cal,:sup)',
                 pid=int(periodo_id), fecha=r['fecha'], suc=r['sucursal'], area=r['area'], cal=r['calificacion'], sup=r['supervisor'])
    conn.close()
    return len(registros)

def procesar_afectaciones(filepath, periodo_id):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    registros = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'Folio': header_found = True; continue
            if not header_found or row[0] is None: continue
            folio, suc, puesto, nomina, nombre, incidencia, fecha, porcentaje, observacion = row[:9]
            registros.append({'folio': str(folio) if folio else '',
                              'sucursal': normalizar_sucursal(suc),
                              'puesto': str(puesto) if puesto else '',
                              'nomina': str(int(nomina)) if nomina and str(nomina) != 'None' else '',
                              'nombre': str(nombre) if nombre else '',
                              'incidencia': str(incidencia) if incidencia else '',
                              'fecha': str(fecha) if fecha else '',
                              'porcentaje': float(porcentaje) if porcentaje else 0,
                              'observacion': str(observacion) if observacion else ''})
    conn = get_db()
    conn.run('DELETE FROM afectaciones WHERE periodo_id=:pid', pid=int(periodo_id))
    for r in registros:
        conn.run('INSERT INTO afectaciones (periodo_id, folio, sucursal, puesto, nomina, nombre, incidencia, fecha, porcentaje, observacion) VALUES (:pid,:folio,:suc,:puesto,:nomina,:nombre,:inc,:fecha,:pct,:obs)',
                 pid=int(periodo_id), folio=r['folio'], suc=r['sucursal'], puesto=r['puesto'],
                 nomina=r['nomina'], nombre=r['nombre'], inc=r['incidencia'],
                 fecha=r['fecha'], pct=r['porcentaje'], obs=r['observacion'])
    conn.close()
    return len(registros)

def procesar_actas(filepath, periodo_id):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    registros = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Año': header_found = True; continue
        if not header_found or row[0] is None: continue
        anio, mes, almacen, area, puesto, nombre, fecha, procedimiento, folio, observaciones = row[:10]
        registros.append({'anio': anio, 'mes': mes, 'almacen': normalizar_sucursal(almacen),
                          'area': str(area) if area else '', 'puesto': str(puesto) if puesto else '',
                          'nombre': str(nombre) if nombre else '', 'fecha': str(fecha) if fecha else '',
                          'procedimiento': str(procedimiento) if procedimiento else '',
                          'folio': str(folio) if folio else '',
                          'observaciones': str(observaciones) if observaciones else '',
                          'porcentaje_afectacion': calcular_afectacion_acta(str(observaciones))})
    conn = get_db()
    conn.run('DELETE FROM actas WHERE periodo_id=:pid', pid=int(periodo_id))
    for r in registros:
        existing = conn.run('SELECT nomina FROM trabajadores WHERE nombre=:n', n=r['nombre'])
        cols = [c['name'] for c in conn.columns]
        nomina = dict(zip(cols, existing[0]))['nomina'] if existing else ''
        conn.run('INSERT INTO actas (periodo_id, anio, mes, almacen, area, puesto, nombre, fecha, procedimiento, folio, observaciones, nomina, porcentaje_afectacion) VALUES (:pid,:anio,:mes,:alm,:area,:puesto,:nombre,:fecha,:proc,:folio,:obs,:nomina,:pct)',
                 pid=int(periodo_id), anio=r['anio'], mes=r['mes'], alm=r['almacen'], area=r['area'],
                 puesto=r['puesto'], nombre=r['nombre'], fecha=r['fecha'], proc=r['procedimiento'],
                 folio=r['folio'], obs=r['observaciones'], nomina=nomina, pct=r['porcentaje_afectacion'])
    conn.close()
    return len(registros)

def procesar_bono_rotulos(filepath, periodo_id):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    registros = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'Sucursal': header_found = True; continue
            if not header_found or row[0] is None or not isinstance(row[0], str): continue
            suc = row[0]
            mat_pop, limp_vis, radio, chequeo, evidencias, total = row[1:7]
            registros.append({'sucursal': suc.strip(),
                              'material_pop': float(mat_pop) if mat_pop else 0,
                              'limpieza_visual': float(limp_vis) if limp_vis else 0,
                              'radio_dpp': float(radio) if radio else 0,
                              'chequeo': float(chequeo) if chequeo else 0,
                              'evidencias': float(evidencias) if evidencias else 0,
                              'total': float(total) if total else 0})
    conn = get_db()
    conn.run('DELETE FROM bono_rotulos WHERE periodo_id=:pid', pid=int(periodo_id))
    for r in registros:
        conn.run('INSERT INTO bono_rotulos (periodo_id, sucursal, material_pop, limpieza_visual, radio_dpp, chequeo, evidencias, total) VALUES (:pid,:suc,:mp,:lv,:rd,:ch,:ev,:tot)',
                 pid=int(periodo_id), suc=r['sucursal'], mp=r['material_pop'], lv=r['limpieza_visual'],
                 rd=r['radio_dpp'], ch=r['chequeo'], ev=r['evidencias'], tot=r['total'])
    conn.close()
    return len(registros)

# =================== CÁLCULO ===================

def get_personal_sucursal(sucursal, conn):
    rows = conn.run('SELECT * FROM trabajadores WHERE sucursal=:s', s=sucursal)
    cols = [c['name'] for c in conn.columns]
    rows = [dict(zip(cols, r)) for r in rows]
    return {
        'encargado_mc': [r for r in rows if r['puesto'] == 'ENCARGADO DE MESA DE CONTROL'],
        'capturista': [r for r in rows if r['puesto'] == 'CAPTURISTA'],
        'rotulista': [r for r in rows if r['puesto'] == 'ROTULISTA'],
        'recibo': [r for r in rows if r['puesto'] == 'RECIBO DE PROVEEDORES'],
        'total': rows
    }

def get_checklist_suc(sucursal, kw, periodo_id, conn):
    rows = conn.run('SELECT calificacion FROM checklists WHERE periodo_id=:pid AND sucursal=:suc AND area ILIKE :kw ORDER BY fecha DESC LIMIT 1',
                    pid=int(periodo_id), suc=sucursal, kw=f'%{kw}%')
    return rows[0][0] if rows else None

def calcular_bono_trabajador(nomina, sucursal, puesto, area, periodo_id):
    conn = get_db()
    resultado = {
        'nomina': nomina, 'sucursal': sucursal, 'puesto': puesto, 'area': area,
        'bono_base': 100.0, 'checklist_aplicado': False, 'checklist_calificacion': None,
        'afectaciones': [], 'actas': [], 'bono_rotulos_externo': None,
        'total_afectaciones': 0.0, 'bono_final': 100.0, 'checklist_heredado': None
    }
    suc_num = normalizar_sucursal(sucursal)
    personal = get_personal_sucursal(suc_num, conn)
    es_unico = len(personal['total']) == 1

    if es_unico:
        cls = []
        for kw in ['MESA', 'ROTUL', 'RECIBO']:
            cal = get_checklist_suc(suc_num, kw, periodo_id, conn)
            if cal is not None: cls.append(float(cal) * 100)
        if cls:
            resultado['checklist_aplicado'] = True
            resultado['checklist_calificacion'] = round(sum(cls)/len(cls), 2)
            resultado['bono_base'] = resultado['checklist_calificacion']

    elif area == 'ROTULOS':
        cal_rot = get_checklist_suc(suc_num, 'ROTUL', periodo_id, conn)
        br_rows = conn.run("SELECT total FROM bono_rotulos WHERE periodo_id=:pid AND sucursal ILIKE :suc LIMIT 1",
                           pid=int(periodo_id), suc=f'%{suc_num}%')
        externo_pct = float(br_rows[0][0]) * 100 if br_rows else 50.0
        resultado['bono_rotulos_externo'] = round(externo_pct, 2)
        if cal_rot is not None:
            checklist_pct = float(cal_rot) * 100
            resultado['checklist_aplicado'] = True
            resultado['checklist_calificacion'] = round(checklist_pct, 2)
            resultado['bono_base'] = round((checklist_pct/100)*50 + externo_pct, 2)
        else:
            # Sin checklist interno: solo calificacion externa ocupa el 50% disponible
            resultado['checklist_aplicado'] = False
            resultado['checklist_calificacion'] = None
            resultado['bono_base'] = round(externo_pct, 2)

    elif area == 'RECIBO':
        cal = get_checklist_suc(suc_num, 'RECIBO', periodo_id, conn)
        if cal is not None:
            resultado['checklist_aplicado'] = True
            resultado['checklist_calificacion'] = round(float(cal)*100, 2)
            resultado['bono_base'] = resultado['checklist_calificacion']

    elif area == 'MESA DE CONTROL':
        if puesto == 'CAPTURISTA':
            if not personal['rotulista']:
                cal_rot = get_checklist_suc(suc_num, 'ROTUL', periodo_id, conn)
                if cal_rot is not None:
                    resultado['checklist_aplicado'] = True
                    resultado['checklist_calificacion'] = round(float(cal_rot)*100, 2)
                    resultado['bono_base'] = resultado['checklist_calificacion']
                    resultado['checklist_heredado'] = 'ROTULOS'
                else:
                    cal_mc = get_checklist_suc(suc_num, 'MESA', periodo_id, conn)
                    if cal_mc is not None:
                        resultado['checklist_aplicado'] = True
                        resultado['checklist_calificacion'] = round(float(cal_mc)*100, 2)
                        resultado['bono_base'] = resultado['checklist_calificacion']
                        resultado['checklist_heredado'] = 'MESA DE CONTROL'
        else:
            cal = get_checklist_suc(suc_num, 'MESA', periodo_id, conn)
            if cal is not None:
                resultado['checklist_aplicado'] = True
                resultado['checklist_calificacion'] = round(float(cal)*100, 2)
                resultado['bono_base'] = resultado['checklist_calificacion']

    # Afectaciones
    afects = conn.run('SELECT folio, fecha, porcentaje, observacion FROM afectaciones WHERE periodo_id=:pid AND nomina=:nom',
                      pid=int(periodo_id), nom=nomina)
    total_afect = 0
    for a in afects:
        pct = float(a[2])
        total_afect += pct
        resultado['afectaciones'].append({'folio': a[0], 'fecha': a[1], 'porcentaje': pct, 'observacion': a[3]})

    # Actas
    trab_rows = conn.run('SELECT nombre FROM trabajadores WHERE nomina=:n', n=nomina)
    nombre_trab = trab_rows[0][0] if trab_rows else ''
    actas = conn.run('SELECT folio, fecha, procedimiento, observaciones, porcentaje_afectacion FROM actas WHERE periodo_id=:pid AND (nomina=:nom OR nombre=:nombre)',
                     pid=int(periodo_id), nom=nomina, nombre=nombre_trab)
    for a in actas:
        pct_acta = float(a[4])
        total_afect += pct_acta
        resultado['actas'].append({'folio': a[0], 'fecha': a[1], 'procedimiento': a[2], 'observaciones': a[3], 'porcentaje': pct_acta})

    resultado['total_afectaciones'] = round(total_afect, 2)
    resultado['bono_final'] = round(max(0, resultado['bono_base'] + total_afect), 2)
    conn.close()
    return resultado

# =================== EXCEL ===================

def generar_excel_reporte(reporte, nombre_periodo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Bonos"
    hf = PatternFill("solid", fgColor="1a2035")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
                    top=Side(style='thin',color='CCCCCC'), bottom=Side(style='thin',color='CCCCCC'))
    ws.merge_cells('A1:K1')
    ws['A1'] = f'REPORTE DE BONOS — {nombre_periodo}'
    ws['A1'].font = Font(bold=True, size=14, color="1a2035")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    headers = ['Nómina','Nombre','Suc.','Nombre Suc.','Puesto','Área','Checklist','Ext. Rótulos','Afectaciones','Bono Final','Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal='center'); cell.border = border
    for ri, t in enumerate(reporte, 3):
        bono = t['bono_final']
        estado = 'EXCELENTE' if bono>=95 else 'BUENO' if bono>=85 else 'REGULAR' if bono>=70 else 'BAJO'
        rfill = PatternFill("solid", fgColor="F8FFF8" if bono>=85 else "FFFFF8" if bono>=70 else "FFF8F8")
        vals = [t['nomina'], t.get('nombre',''), t['sucursal'], t.get('nombre_suc',''), t['puesto'], t['area'],
                f"{t['checklist_calificacion']}%" if t['checklist_calificacion'] else '100% (base)',
                f"{t['bono_rotulos_externo']}%" if t['bono_rotulos_externo'] is not None else '—',
                f"{t['total_afectaciones']}%", f"{bono}%", estado]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.fill = rfill
            cell.alignment = Alignment(horizontal='center' if col>5 else 'left')
            cell.border = border
    ws2 = wb.create_sheet("Detalle Afectaciones")
    dh = ['Nómina','Nombre','Sucursal','Tipo','Folio','Fecha','% Afectación','Descripción']
    for col, h in enumerate(dh, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = hf; cell.font = hfont; cell.border = border
    rd = 2
    for t in reporte:
        for a in t.get('afectaciones', []):
            for col, val in enumerate([t['nomina'],t.get('nombre',''),t['sucursal'],'Afectación',a['folio'],a['fecha'],f"{a['porcentaje']}%",a['observacion']],1):
                ws2.cell(row=rd, column=col, value=val)
            rd += 1
        for a in t.get('actas', []):
            for col, val in enumerate([t['nomina'],t.get('nombre',''),t['sucursal'],'Acta',a['folio'],a['fecha'],f"{a['porcentaje']}%",a['observaciones']],1):
                ws2.cell(row=rd, column=col, value=val)
            rd += 1
    for wso in [ws, ws2]:
        for col in wso.columns:
            mx = max((len(str(cell.value or '')) for cell in col), default=0)
            wso.column_dimensions[col[0].column_letter].width = min(mx+4, 50)
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return output

# =================== RUTAS ===================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/periodos', methods=['GET'])
def get_periodos():
    conn = get_db()
    rows = conn.run('SELECT id, nombre, mes, anio, fecha_inicio, fecha_fin FROM periodos ORDER BY id DESC')
    conn.close()
    keys = ['id','nombre','mes','anio','fecha_inicio','fecha_fin']
    return jsonify([dict(zip(keys, r)) for r in rows])

@app.route('/api/periodos', methods=['POST'])
def crear_periodo():
    data = request.json
    conn = get_db()
    try:
        rows = conn.run('INSERT INTO periodos (nombre, mes, anio, fecha_inicio, fecha_fin) VALUES (:n,:m,:a,:fi,:ff) RETURNING id',
                        n=data['nombre'], m=data['mes'], a=data['anio'],
                        fi=data.get('fecha_inicio',''), ff=data.get('fecha_fin',''))
        pid = rows[0][0]
        conn.close()
        return jsonify({'success': True, 'id': pid})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file'})
    file = request.files['file']
    tipo = request.form.get('tipo')
    periodo_id = request.form.get('periodo_id')
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(UPLOAD_DIR, file.filename)
    file.save(filepath)
    try:
        if tipo == 'checklist': n = procesar_checklist(filepath, periodo_id); return jsonify({'success': True, 'registros': n})
        elif tipo == 'afectaciones': n = procesar_afectaciones(filepath, periodo_id); return jsonify({'success': True, 'registros': n})
        elif tipo == 'actas': n = procesar_actas(filepath, periodo_id); return jsonify({'success': True, 'registros': n})
        elif tipo == 'rotulos': n = procesar_bono_rotulos(filepath, periodo_id); return jsonify({'success': True, 'registros': n})
        elif tipo == 'catalogo':
            ins, act = cargar_catalogo(filepath)
            return jsonify({'success': True, 'registros': ins+act, 'insertados': ins, 'actualizados': act})
        else: return jsonify({'success': False, 'error': 'Tipo no reconocido'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/trabajadores', methods=['GET'])
def get_trabajadores():
    conn = get_db()
    q = request.args.get('q', '')
    if q:
        rows = conn.run("SELECT id, nomina, nombre, sucursal, puesto, area, nombre_suc FROM trabajadores WHERE nombre ILIKE :q OR nomina ILIKE :q OR sucursal ILIKE :q ORDER BY sucursal::integer, nombre", q=f'%{q}%')
    else:
        rows = conn.run("SELECT id, nomina, nombre, sucursal, puesto, area, nombre_suc FROM trabajadores ORDER BY sucursal::integer, nombre")
    conn.close()
    keys = ['id','nomina','nombre','sucursal','puesto','area','nombre_suc']
    return jsonify([dict(zip(keys, r)) for r in rows])

@app.route('/api/trabajadores', methods=['POST'])
def agregar_trabajador():
    data = request.json
    conn = get_db()
    try:
        conn.run('INSERT INTO trabajadores (nomina, nombre, sucursal, puesto, area, nombre_suc) VALUES (:nomina,:nombre,:suc,:puesto,:area,:nsuc) ON CONFLICT (nomina) DO UPDATE SET nombre=EXCLUDED.nombre, sucursal=EXCLUDED.sucursal, puesto=EXCLUDED.puesto, area=EXCLUDED.area, nombre_suc=EXCLUDED.nombre_suc',
                 nomina=data['nomina'], nombre=data['nombre'], suc=data['sucursal'],
                 puesto=data['puesto'], area=data['area'], nsuc=data.get('nombre_suc',''))
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/trabajadores/<int:id>', methods=['DELETE'])
def eliminar_trabajador(id):
    conn = get_db()
    conn.run('DELETE FROM trabajadores WHERE id=:id', id=id)
    conn.close()
    return jsonify({'success': True})

def get_reporte_data(periodo_id, sucursal='', area='', buscar=''):
    conn = get_db()
    query = "SELECT id, nomina, nombre, sucursal, puesto, area, nombre_suc FROM trabajadores WHERE 1=1"
    params = {}
    if sucursal: query += ' AND sucursal=:suc'; params['suc'] = sucursal
    if area: query += ' AND area=:area'; params['area'] = area
    if buscar: query += ' AND (nombre ILIKE :b OR nomina ILIKE :b)'; params['b'] = f'%{buscar}%'
    query += ' ORDER BY sucursal::integer, nombre'
    rows = conn.run(query, **params)
    conn.close()
    keys = ['id','nomina','nombre','sucursal','puesto','area','nombre_suc']
    trabajadores = [dict(zip(keys, r)) for r in rows]
    reporte = []
    for t in trabajadores:
        bono = calcular_bono_trabajador(t['nomina'], t['sucursal'], t['puesto'], t['area'], periodo_id)
        bono['nombre'] = t['nombre']
        bono['nombre_suc'] = t.get('nombre_suc', '')
        reporte.append(bono)
    return reporte

@app.route('/api/reporte', methods=['GET'])
def get_reporte():
    periodo_id = request.args.get('periodo_id')
    if not periodo_id: return jsonify({'error': 'Periodo requerido'})
    reporte = get_reporte_data(periodo_id, request.args.get('sucursal',''), request.args.get('area',''), request.args.get('buscar',''))
    return jsonify(reporte)

@app.route('/api/reporte/excel', methods=['GET'])
def exportar_excel():
    periodo_id = request.args.get('periodo_id')
    if not periodo_id: return jsonify({'error': 'Periodo requerido'})
    conn = get_db()
    rows = conn.run('SELECT nombre FROM periodos WHERE id=:id', id=int(periodo_id))
    nombre_periodo = rows[0][0] if rows else 'Periodo'
    conn.close()
    reporte = get_reporte_data(periodo_id, request.args.get('sucursal',''), request.args.get('area',''), request.args.get('buscar',''))
    excel = generar_excel_reporte(reporte, nombre_periodo)
    tmp_path = os.path.join(UPLOAD_DIR, f'reporte_tmp.xlsx')
    with open(tmp_path, 'wb') as f:
        f.write(excel.read())
    return send_file(tmp_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Bonos_{nombre_periodo.replace(" ","_")}.xlsx')

@app.route('/api/sucursales', methods=['GET'])
def get_sucursales():
    conn = get_db()
    rows = conn.run('SELECT DISTINCT sucursal, nombre_suc FROM trabajadores ORDER BY sucursal::integer')
    conn.close()
    return jsonify([{'num': r[0], 'nombre': r[1]} for r in rows])

os.makedirs(UPLOAD_DIR, exist_ok=True)
try:
    init_db()
except Exception as e:
    print(f"DB init warning: {e}")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
